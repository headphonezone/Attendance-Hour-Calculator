import io
import re
from datetime import datetime, date
from collections import defaultdict
import calendar

import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

C_HEADER_BG  = "1F4E79"
C_HEADER_FG  = "FFFFFF"
C_SUBHDR_BG  = "2E75B6"
C_EXCESS_BG  = "E2EFDA"
C_SHORT_BG   = "FCE4D6"
C_NEUTRAL_BG = "DEEAF1"
C_ALT_ROW    = "F2F7FB"
C_BORDER     = "BDD7EE"
C_HOLIDAY_BG = "FFF2CC"
C_WFH_BG     = "E2F0D9"
C_RAW_HDR    = "375623"

def make_border(color=C_BORDER):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def make_font(bold=False, color="000000", size=10):
    return Font(name='Arial', bold=bold, color=color, size=size)

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_align(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def style(cell, *, bold=False, fg="000000", size=10,
          fill=C_ALT_ROW, align_h='center', wrap=False, border=True):
    cell.font      = make_font(bold, fg, size)
    cell.fill      = make_fill(fill)
    cell.alignment = make_align(align_h, wrap=wrap)
    if border:
        cell.border = make_border()

def parse_punches(cell_value):
    if not cell_value:
        return []
    return [t.strip() for t in str(cell_value).split('\n') if t.strip()]

def minutes_to_hhmm(total_minutes):
    h, m = int(total_minutes // 60), int(total_minutes % 60)
    return f"{h}.{m:02d}"

def decimal_to_hhmm(decimal_hours):
    return minutes_to_hhmm(round(decimal_hours * 60))

# Excel stores time as a fraction of a 24-hour day. To show real hour
# totals (which routinely exceed 24) as proper H:MM — 60 minutes to the
# hour, not 100 — we store hours/24 and format the cell as elapsed time.
# Excel cannot render a negative duration in this format at all (always
# shows ####), so any value that can go negative (e.g. Net Hours) must be
# built as a text label via TEXT()/"-" concatenation instead.
TIME_FMT = "[h]:mm"

def to_excel_time(decimal_hours):
    return round(decimal_hours, 4) / 24

def compute_hours_from_pair(t_in_str, t_out_str):
    try:
        t_in  = datetime.strptime(t_in_str,  "%H:%M")
        t_out = datetime.strptime(t_out_str, "%H:%M")
        diff  = int((t_out - t_in).total_seconds() / 60)
        if diff <= 0:
            return 0.0, "0.00"
        return round(diff / 60, 4), minutes_to_hhmm(diff)
    except ValueError:
        return 0.0, "0.00"

def get_week_number(day, year, month):
    wc, fw = 1, date(year, month, 1).weekday()
    for d in range(1, day + 1):
        if date(year, month, d).weekday() == 0 and d != 1:
            if not (d == 2 and fw == 6):
                wc += 1
    return wc

def get_week_target(relative_wk, year, month, daily_target):
    wd = 0
    for d_int in range(1, 32):
        try:
            if get_week_number(d_int, year, month) == relative_wk:
                wd += 1
        except ValueError:
            break
    return round(wd * daily_target, 2)

def get_month_sundays(year, month):
    total = calendar.monthrange(year, month)[1]
    return [d for d in range(1, total + 1) if date(year, month, d).weekday() == 6]

# ── ID normalization ────────────────────────────────────────────────────
# Excel stores IDs as numbers unless the cell is formatted as text, so the
# same employee ID can round-trip as "91", "91.0", " 91 " or "ABC01" vs
# "abc01" depending on which sheet it came from. Salary matching is by ID,
# so every ID must funnel through this before being stored or compared.
def normalize_id(val):
    if val is None:
        return ""
    s = str(val).strip()
    if s.endswith(".0"):
        try:
            f = float(s)
            if f == int(f):
                s = str(int(f))
        except ValueError:
            pass
    # Purely numeric IDs: strip leading zeros so "007" (kept as text in one
    # sheet) matches "7" (typed as a plain number in the other) — Excel
    # can't preserve leading zeros in a numeric cell, so the padded form
    # only ever shows up on one side.
    if s.isdigit():
        s = str(int(s))
    return s.upper()

# ── Salary Master parsing ──────────────────────────────────────────────
# Reusable file (ID, Name, Salary columns, header row optional) so the
# user only maintains one small sheet and re-uploads it every month
# instead of retyping salaries for every employee each time.
def _parse_salary_number(val):
    """Coerce a salary cell to a float. Handles plain numbers as well as
    text values with commas/currency symbols (e.g. "15,000", "Rs. 15000")
    that a straight float() call would reject."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    s = re.sub(r"[^\d.\-]", "", s)
    if not s or s in ("-", "."):
        return None
    try:
        return float(s)
    except ValueError:
        return None

def parse_salary_file(uploaded_file):
    salary_map, name_map, skipped = {}, {}, []
    fname = (uploaded_file.name or "").lower()
    try:
        if fname.endswith(".csv"):
            import csv as _csv
            content = uploaded_file.getvalue().decode("utf-8-sig")
            reader  = _csv.reader(content.splitlines())
            rows    = list(reader)
        else:
            # data_only=True: if a Salary cell holds a formula (e.g. a rate
            # lookup), read its last-saved computed value instead of the
            # formula text, which would otherwise fail number parsing.
            wb   = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
            ws   = wb[wb.sheetnames[0]]
            rows = list(ws.iter_rows(values_only=True))

        for row_num, row in enumerate(rows, 1):
            if not row or all(c is None for c in row):
                continue
            row = list(row) + [None] * max(0, 3 - len(row))
            emp_id_raw, emp_name, salary_raw = row[0], row[1], row[2]

            if emp_id_raw is None:
                skipped.append((row_num, "—", "missing ID"))
                continue
            emp_id = normalize_id(emp_id_raw)
            if not emp_id or emp_id.lower() in ("id", "employee id"):
                continue

            salary = _parse_salary_number(salary_raw)
            if salary is None:
                skipped.append((row_num, emp_id, f"unreadable salary value: {salary_raw!r}"))
                continue
            if salary <= 0:
                skipped.append((row_num, emp_id, "salary is 0 or negative"))
                continue

            salary_map[emp_id] = salary
            name_map[emp_id]   = str(emp_name).strip() if emp_name else ""
    except Exception as e:
        skipped.append(("?", "?", f"file read error: {e}"))
    return salary_map, name_map, skipped

def make_salary_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Salary Master"
    ws.append(["ID", "Name", "Monthly Salary"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── Part-Time Master parsing ────────────────────────────────────────────
# Reusable file (ID, Name columns) so part-time employees don't need to be
# re-selected from a dropdown every month — re-upload the same file.
def parse_id_name_file(uploaded_file):
    id_set, name_map, skipped = set(), {}, []
    fname = (uploaded_file.name or "").lower()
    try:
        if fname.endswith(".csv"):
            import csv as _csv
            content = uploaded_file.getvalue().decode("utf-8-sig")
            reader  = _csv.reader(content.splitlines())
            rows    = list(reader)
        else:
            wb   = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
            ws   = wb[wb.sheetnames[0]]
            rows = list(ws.iter_rows(values_only=True))

        for row_num, row in enumerate(rows, 1):
            if not row or all(c is None for c in row):
                continue
            row = list(row) + [None] * max(0, 2 - len(row))
            emp_id_raw, emp_name = row[0], row[1]

            if emp_id_raw is None:
                skipped.append((row_num, "—", "missing ID"))
                continue
            emp_id = normalize_id(emp_id_raw)
            if not emp_id or emp_id.lower() in ("id", "employee id"):
                continue

            id_set.add(emp_id)
            name_map[emp_id] = str(emp_name).strip() if emp_name else ""
    except Exception as e:
        skipped.append(("?", "?", f"file read error: {e}"))
    return id_set, name_map, skipped

def make_parttime_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Part-Time Master"
    ws.append(["ID", "Name"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def parse_logs_sheet(ws):
    all_rows   = list(ws.iter_rows(values_only=True))
    period_str = ""
    year, month = datetime.now().year, datetime.now().month

    for row in all_rows[:5]:
        for cell in row:
            if cell and isinstance(cell, str) and '~' in cell:
                period_str = cell.strip()
                try:
                    dt          = datetime.strptime(period_str.split('~')[0].strip(), "%Y/%m/%d")
                    year, month = dt.year, dt.month
                except Exception:
                    pass

    raw_records, emp_order = {}, []
    i = 0
    while i < len(all_rows):
        row = all_rows[i]
        if row and row[0] == 'No :':
            emp_no    = normalize_id(row[2]) if row[2] else 'Unknown'
            emp_name  = str(row[10]).strip() if row[10] else 'Unnamed'
            uid       = f"{emp_name.title()} (ID: {emp_no})"
            days_row  = all_rows[i - 1] if i > 0 else []
            pr_idx    = i + 1
            if pr_idx < len(all_rows):
                punch_row   = all_rows[pr_idx]
                day_punches = {}
                for col, day_num in enumerate(days_row):
                    if not isinstance(day_num, int): continue
                    if col >= len(punch_row): continue
                    try:
                        date(year, month, day_num)
                    except:
                        continue
                    punches = parse_punches(punch_row[col])
                    if punches:
                        day_punches[day_num] = punches
                if uid not in raw_records:
                    raw_records[uid] = {'name': emp_name, 'id': emp_no, 'punches': {}}
                    emp_order.append(uid)
                raw_records[uid]['punches'].update(day_punches)
        i += 1
    return raw_records, emp_order, period_str, year, month

# ── CHANGE 1 + 2: Skip relieved employees + inject paid holiday hrs ──────
def build_employees_dec(emp_order, raw_records, fixes, wfh_records, year, month,
                         holiday_dates, part_time_list=None, pt_daily_target=4.0,
                         daily_target=8.5):
    employees_dec    = {}
    holiday_day_nums = set(hd.day for hd in holiday_dates if hd.year == year and hd.month == month)
    month_sundays    = set(get_month_sundays(year, month))
    part_time_list    = part_time_list or []

    for uid in emp_order:
        p_dict  = raw_records[uid]['punches']
        has_wfh = bool(wfh_records.get(uid, {}))
        emp_daily_target = pt_daily_target if uid in part_time_list else daily_target

        # Skip relieved employees — no punches and no WFH for the entire month
        if not p_dict and not has_wfh:
            continue

        f_dict    = fixes.get(uid, {})
        week_data = defaultdict(dict)

        for day, p in p_dict.items():
            if day in f_dict:
                dec, _ = compute_hours_from_pair(f_dict[day]['in'], f_dict[day]['out'])
            elif len(p) >= 2:
                dec = sum(compute_hours_from_pair(p[i], p[i+1])[0] for i in range(0, len(p)-1, 2))
            elif len(p) == 1:
                h   = int(p[0].split(':')[0]) if ':' in p[0] else 0
                dec, _ = (compute_hours_from_pair("09:30", p[0]) if h >= 12
                          else compute_hours_from_pair(p[0], "18:00"))
            else:
                dec = 0.0
            if dec > 0:
                week_data[get_week_number(day, year, month)][day] = dec

        # Inject WFH hours for days with no punch data
        for day, info in wfh_records.get(uid, {}).items():
            hrs = info.get('hours', 0.0)
            if hrs > 0:
                wk = get_week_number(day, year, month)
                if day not in week_data.get(wk, {}):
                    week_data[wk][day] = hrs

        # Inject paid holiday hrs (at the employee's own daily target) —
        # credited to hours but NOT a working day. Applies to Sundays too.
        # Only fills days with no punch/WFH entry.
        for hday in holiday_day_nums:
            try:
                date(year, month, hday)
            except:
                continue
            wk = get_week_number(hday, year, month)
            if hday not in week_data.get(wk, {}):
                week_data[wk][hday] = emp_daily_target   # paid holiday

        # Auto-credit Sundays with no actual punch/WFH data — treated like a
        # paid day off (same as a holiday) so they never register as shortage.
        # Real Sunday punches (already added above) are left untouched.
        for sday in month_sundays:
            if sday in holiday_day_nums:
                continue   # already handled by holiday injection
            wk = get_week_number(sday, year, month)
            if sday not in week_data.get(wk, {}):
                week_data[wk][sday] = emp_daily_target

        if week_data:
            employees_dec[uid] = dict(week_data)
    return employees_dec

def get_leave_days(uid, raw_records, year, month, holiday_dates, wfh_records):
    total_days   = calendar.monthrange(year, month)[1]
    punched_days = set(raw_records[uid]['punches'].keys())
    holiday_nums = set(hd.day for hd in holiday_dates if hd.year == year and hd.month == month)
    wfh_days     = set(wfh_records.get(uid, {}).keys())
    sundays      = set(get_month_sundays(year, month))
    leave = 0
    for d in range(1, total_days + 1):
        # Sundays are auto-credited like a paid day off, so an unpunched
        # Sunday is not a leave day.
        if d in holiday_nums or d in wfh_days or d in sundays: continue
        if d not in punched_days:
            leave += 1
    return leave

def get_leave_days_by_week(uid, raw_records, year, month, holiday_dates, wfh_records):
    """Same definition of 'leave day' as get_leave_days, grouped by week
    number, so a week's target can be reduced by the leave taken in it —
    a leave day should count only as leave, not also as shortage."""
    total_days   = calendar.monthrange(year, month)[1]
    punched_days = set(raw_records[uid]['punches'].keys())
    holiday_nums = set(hd.day for hd in holiday_dates if hd.year == year and hd.month == month)
    wfh_days     = set(wfh_records.get(uid, {}).keys())
    sundays      = set(get_month_sundays(year, month))
    by_week = defaultdict(int)
    for d in range(1, total_days + 1):
        if d in holiday_nums or d in wfh_days or d in sundays: continue
        if d not in punched_days:
            by_week[get_week_number(d, year, month)] += 1
    return dict(by_week)

def get_effective_week_target(wk, year, month, daily_target, leave_by_week):
    """Weekly target minus the daily target for each leave day taken that
    week, so leave is reflected once (as leave) instead of twice (leave
    and shortage for the same day)."""
    base       = get_week_target(wk, year, month, daily_target)
    leave_days = leave_by_week.get(wk, 0)
    return round(max(0.0, base - leave_days * daily_target), 2)

def get_holidays_on_leave(uid, raw_records, year, month, holiday_dates, wfh_records):
    punched_days = set(raw_records[uid]['punches'].keys())
    wfh_days     = set(wfh_records.get(uid, {}).keys())
    count = 0
    for hd in holiday_dates:
        if hd.year != year or hd.month != month:
            continue
        d = hd.day
        if d in wfh_days:   continue
        if d not in punched_days:
            count += 1
    return count

# ── CHANGE 2b: get_days_worked — holidays & auto-credited Sundays NOT counted
# as working days. Based on actual raw punch/WFH data, not injected hours,
# so an auto-credited (unpunched) Sunday never counts as a day worked —
# only a Sunday with a real punch does.
def get_days_worked(uid, raw_records, wfh_records, holiday_dates, year, month):
    holiday_day_nums = set(hd.day for hd in holiday_dates if hd.year == year and hd.month == month)
    punched_days     = {d for d, p in raw_records[uid]['punches'].items() if p} - holiday_day_nums
    wfh_non_holiday  = {d for d in wfh_records.get(uid, {}) if d not in holiday_day_nums}
    return len(punched_days | wfh_non_holiday)

def sum_week_hours(day_dict):
    return sum(day_dict.values())

RAW_SHEET          = "_RawData"
RAW_DATA_START_ROW = 2

def write_raw_data_sheet(wb, employees_dec, emp_order, raw_records,
                          year, month, daily_target, part_time_list, pt_daily_target,
                          period_str, holiday_dates, wfh_records):
    ws = wb.create_sheet(RAW_SHEET)
    ws.sheet_state = 'hidden'

    for col, h in enumerate(["UID_KEY", "ID", "Name", "Week",
                              "HoursWorked", "Target", "Excess", "Shortage"], 1):
        ws.cell(row=1, column=col, value=h)

    row     = RAW_DATA_START_ROW
    row_map = {}

    for uid in emp_order:
        if uid not in employees_dec:
            continue
        current_daily = pt_daily_target if uid in part_time_list else daily_target
        week_dict     = employees_dec[uid]
        row_map[uid]  = {}
        leave_by_week = get_leave_days_by_week(uid, raw_records, year, month,
                                                holiday_dates, wfh_records)

        for wk in sorted(week_dict.keys()):
            wk_target  = get_effective_week_target(wk, year, month, current_daily, leave_by_week)
            wk_hrs_dec = sum_week_hours(week_dict[wk])

            ws.cell(row=row, column=1, value=uid)
            ws.cell(row=row, column=2, value=raw_records[uid]['id'])
            ws.cell(row=row, column=3, value=raw_records[uid]['name'].title())
            ws.cell(row=row, column=4, value=f"Week {wk}")
            ws.cell(row=row, column=5, value=to_excel_time(wk_hrs_dec)).number_format = TIME_FMT
            ws.cell(row=row, column=6, value=to_excel_time(wk_target)).number_format  = TIME_FMT
            ws.cell(row=row, column=7, value=f"=MAX(0,E{row}-F{row})").number_format  = TIME_FMT
            ws.cell(row=row, column=8, value=f"=MAX(0,F{row}-E{row})").number_format  = TIME_FMT

            row_map[uid][wk] = row
            row += 1

    return ws, row_map

def write_summary_sheet(wb, employees_dec, emp_order, raw_records,
                         period_str, year, month, daily_target, part_time_list,
                         pt_daily_target, row_map, holiday_dates, wfh_records):
    ws = wb.create_sheet("Weekly Summary")

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = f"Weekly Attendance Summary | {period_str}"
    c.font      = make_font(True, C_HEADER_FG, 14)
    c.fill      = make_fill(C_HEADER_BG)
    c.alignment = make_align()

    ws.merge_cells("A2:H2")
    note = ws["A2"]
    note.value     = "⚠️  Edit 'Hours Worked' values here (as H:MM, e.g. 8:30) — Consolidated Report updates automatically via formulas."
    note.font      = make_font(False, "7B3F00", 9)
    note.fill      = make_fill(C_HOLIDAY_BG)
    note.alignment = make_align()

    headers = ["ID", "Employee Name", "Week",
               "Hours Worked ✏️", "Target Hours", "Excess", "Shortage", "Status"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = (
            make_font(True, C_HEADER_FG), make_fill(C_SUBHDR_BG), make_align(), make_border()
        )

    row = 5
    for uid in emp_order:
        if uid not in employees_dec or uid not in row_map:
            continue
        current_daily = pt_daily_target if uid in part_time_list else daily_target
        leave_by_week = get_leave_days_by_week(uid, raw_records, year, month,
                                                holiday_dates, wfh_records)
        for wk in sorted(employees_dec[uid].keys()):
            raw_row = row_map[uid].get(wk)
            if raw_row is None:
                continue

            wk_hrs_dec    = sum_week_hours(employees_dec[uid][wk])
            wk_target     = get_effective_week_target(wk, year, month, current_daily, leave_by_week)

            ws.cell(row=row, column=1, value=raw_records[uid]['id'])
            ws.cell(row=row, column=2, value=raw_records[uid]['name'].title())
            ws.cell(row=row, column=3, value=f"Week {wk}")
            ws.cell(row=row, column=4, value=to_excel_time(wk_hrs_dec)).number_format = TIME_FMT
            ws.cell(row=row, column=5, value=to_excel_time(wk_target)).number_format  = TIME_FMT
            ws.cell(row=row, column=6, value=f"=MAX(0,D{row}-E{row})").number_format  = TIME_FMT
            ws.cell(row=row, column=7, value=f"=MAX(0,E{row}-D{row})").number_format  = TIME_FMT
            ws.cell(row=row, column=8,
                    value=f'=IF(D{row}>E{row},"EXCESS",IF(D{row}<E{row},"SHORTAGE","ON TARGET"))')

            row_map[uid][wk] = (raw_row, row)

            if wk_hrs_dec > wk_target:
                fill_c = C_EXCESS_BG
            elif wk_hrs_dec < wk_target:
                fill_c = C_SHORT_BG
            else:
                fill_c = C_ALT_ROW

            for col in range(1, 9):
                c = ws.cell(row=row, column=col)
                c.fill      = make_fill(fill_c)
                c.border    = make_border()
                c.alignment = make_align()
                c.font      = make_font()
                if col == 4:
                    c.font = make_font(bold=True)

            row += 1

    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 14
    for ltr in ['F', 'G', 'H']:
        ws.column_dimensions[ltr].width = 14

    return ws, row_map

# ── CHANGE 3: Net split into Net Hours (number) + Status (label) ──────────────
def write_consolidated_sheet(wb, employees_dec, emp_order, raw_records, period_str,
                              year, month, daily_target, part_time_list, pt_daily_target,
                              holiday_dates, wfh_records, row_map, salary_map=None):
    ws = wb.create_sheet("Consolidated Report")
    salary_map = salary_map or {}

    headers = [
        "ID", "Employee Name",
        "Total Hours", "Total Target", "Total Excess", "Total Shortage",
        "Net Hours",    # G — number only (positive=excess, negative=shortage)
        "Status",       # H — "Excess" / "Shortage" / "On Target"
        "Days Worked", "Leave Days", "Holidays on Leave", "Sundays", "Holidays",
        "Monthly Salary", "Calculated Salary"
    ]
    num_cols = len(headers)

    ws.merge_cells(f"A1:{get_column_letter(num_cols)}1")
    c = ws["A1"]
    c.value     = f"Total Monthly Consolidation | {period_str}"
    c.font      = make_font(True, C_HEADER_FG, 14)
    c.fill      = make_fill(C_HEADER_BG)
    c.alignment = make_align()

    hdr_row = 3
    if holiday_dates:
        ws.merge_cells(f"A2:{get_column_letter(num_cols)}2")
        ws["A2"].value     = "Holidays: " + ", ".join(hd.strftime("%d-%b-%Y") for hd in sorted(holiday_dates))
        ws["A2"].font      = make_font(True, "7B3F00", 9)
        ws["A2"].fill      = make_fill(C_HOLIDAY_BG)
        ws["A2"].alignment = make_align()
        hdr_row = 4

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=hdr_row, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = (
            make_font(True, C_HEADER_FG), make_fill(C_SUBHDR_BG), make_align(), make_border()
        )

    num_sundays  = len(get_month_sundays(year, month))
    num_holidays = len([hd for hd in holiday_dates if hd.year == year and hd.month == month])

    total_summary_rows = sum(len(d) for d in employees_dec.values())
    sum_end_row        = max(5, 5 + total_summary_rows - 1)
    ws_ref             = "'Weekly Summary'"

    data_row = hdr_row + 1

    for uid in emp_order:
        if uid not in employees_dec:
            continue

        emp_name_title = raw_records[uid]['name'].title()
        emp_id         = raw_records[uid]['id']

        name_col     = f"{ws_ref}!$B$5:$B${sum_end_row}"
        hours_col    = f"{ws_ref}!$D$5:$D${sum_end_row}"
        target_col   = f"{ws_ref}!$E$5:$E${sum_end_row}"
        excess_col   = f"{ws_ref}!$F$5:$F${sum_end_row}"
        shortage_col = f"{ws_ref}!$G$5:$G${sum_end_row}"
        crit         = f'"{emp_name_title}"'

        # Values here are day-fractions (hours/24) so cells can be formatted
        # as real elapsed time ([h]:mm = 60 min/hr) instead of base-10
        # decimal. Rounding to 4 dp keeps ~0.35-second precision.
        #
        # Total Hours is capped at each week's target (excess excluded) —
        # actual hours worked minus the excess portion, so a week with
        # overtime doesn't inflate this figure. Total Excess (col E) still
        # shows the excess separately. Total Hours + Total Excess always
        # recovers the true raw hours worked, since Excess is never
        # negative — unlike Net Hours, which can be negative and would
        # double-subtract an already-reflected shortage if added here.
        f_hrs      = (f"=ROUND(SUMIF({name_col},{crit},{hours_col})"
                       f"-MAX(0,SUMIF({name_col},{crit},{excess_col})),4)")
        f_target   = f"=ROUND(SUMIF({name_col},{crit},{target_col}),4)"
        f_excess   = f"=ROUND(MAX(0,SUMIF({name_col},{crit},{excess_col})),4)"
        f_shortage = f"=ROUND(MAX(0,SUMIF({name_col},{crit},{shortage_col})),4)"

        # G: Net Hours — Excel can't render a negative [h]:mm duration
        # (always shows ####), so build the label as text instead: positive
        # net formats normally, negative net gets a "-" prefix on the
        # absolute difference.
        e_col = get_column_letter(5)
        f_col = get_column_letter(6)
        net_diff  = f"({e_col}{data_row}-{f_col}{data_row})"
        f_net_text = (
            f'=IF({net_diff}>=0,TEXT({net_diff},"[h]:mm"),'
            f'"-"&TEXT(-{net_diff},"[h]:mm"))'
        )

        # H: Status label — based on the same excess/shortage difference
        f_status = (
            f'=IF({net_diff}>0,"Excess",'
            f'IF({net_diff}<0,"Shortage","On Target"))'
        )

        days_worked       = get_days_worked(uid, raw_records, wfh_records, holiday_dates, year, month)
        leave_days        = get_leave_days(uid, raw_records, year, month, holiday_dates, wfh_records)
        holidays_on_leave = get_holidays_on_leave(uid, raw_records, year, month, holiday_dates, wfh_records)

        wk_dict          = employees_dec[uid]
        current_daily     = pt_daily_target if uid in part_time_list else daily_target
        leave_by_week     = get_leave_days_by_week(uid, raw_records, year, month,
                                                     holiday_dates, wfh_records)
        total_hours_dec   = sum(sum(d.values()) for d in wk_dict.values())
        total_target_dec  = sum(get_effective_week_target(wk, year, month, current_daily, leave_by_week)
                                 for wk in wk_dict)
        net               = round(total_hours_dec - total_target_dec, 2)

        # Salary: per_hour = Monthly Salary / Total Target hours;
        # Calculated Salary = per_hour * Total Hours Worked. This alone
        # already prorates for excess/shortage relative to target — adding
        # Net Hours on top would double-count the deviation and could go
        # negative when shortage exceeds half of hours worked.
        monthly_salary = salary_map.get(normalize_id(emp_id))
        if monthly_salary and total_target_dec > 0:
            per_hour           = monthly_salary / total_target_dec
            calculated_salary  = round(per_hour * total_hours_dec, 2)
        else:
            calculated_salary  = None

        vals = [
            emp_id,             # A — 1
            emp_name_title,     # B — 2
            f_hrs,              # C — 3
            f_target,           # D — 4
            f_excess,           # E — 5
            f_shortage,         # F — 6
            f_net_text,         # G — 7  Net Hours (text label, e.g. "-8:30")
            f_status,           # H — 8  Status label
            days_worked,        # I — 9
            leave_days,         # J — 10
            holidays_on_leave,  # K — 11
            num_sundays,        # L — 12
            num_holidays,       # M — 13
            monthly_salary,     # N — 14
            calculated_salary,  # O — 15
        ]

        for col, v in enumerate(vals, 1):
            c = ws.cell(row=data_row, column=col, value=v)
            if col in (3, 4, 5, 6):
                c.number_format = TIME_FMT
            if col in (14, 15):
                c.number_format = "#,##0.00"
            if col in (7, 8):
                fill_c = C_EXCESS_BG if net > 0 else (C_SHORT_BG if net < 0 else C_ALT_ROW)
            elif col == 11 and isinstance(v, int) and v > 0:
                fill_c = C_HOLIDAY_BG
            elif col == 15 and calculated_salary is None:
                fill_c = C_SHORT_BG
            else:
                fill_c = C_ALT_ROW
            c.fill, c.border, c.alignment, c.font = (
                make_fill(fill_c), make_border(), make_align(), make_font()
            )

        data_row += 1

    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['K'].width = 18
    for ltr in ['C', 'D', 'E', 'F']:
        ws.column_dimensions[ltr].width = 15
    for ltr in ['I', 'J', 'L', 'M']:
        ws.column_dimensions[ltr].width = 13
    for ltr in ['N', 'O']:
        ws.column_dimensions[ltr].width = 16

def write_individual_sheet(wb, uid, week_dict, period_str, year, month,
                           daily_target, is_part_time, pt_daily_target,
                           holiday_dates, wfh_records, raw_records):
    ws_name = (uid[:28]
               .replace(":", "").replace("/", "").replace("*", "")
               .replace("?", "").replace("[", "").replace("]", "").strip())
    ws = wb.create_sheet(ws_name)

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value     = f"Attendance Details | {uid} {'(Part-Time)' if is_part_time else ''}"
    c.font      = make_font(True, C_HEADER_FG, 12)
    c.fill      = make_fill(C_HEADER_BG)
    c.alignment = make_align()

    holiday_day_nums = set(hd.day for hd in holiday_dates if hd.year == year and hd.month == month)
    wfh_dict         = wfh_records.get(uid, {})
    punched_days     = set(raw_records[uid]['punches'].keys())
    current_daily    = pt_daily_target if is_part_time else daily_target
    leave_by_week    = get_leave_days_by_week(uid, raw_records, year, month,
                                               holiday_dates, wfh_records)

    row = 3
    for wk in sorted(week_dict.keys()):
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1, value=f"WEEK {wk}")
        c.font, c.fill, c.alignment = make_font(True, C_HEADER_FG), make_fill(C_SUBHDR_BG), make_align()
        row += 1

        for col, h in enumerate(["Date", "Day", "Hours Worked", "Note"], 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font, c.fill, c.alignment, c.border = (
                make_font(True, C_HEADER_FG), make_fill(C_SUBHDR_BG), make_align(), make_border()
            )
        row += 1

        for day, hrs in sorted(week_dict[wk].items()):
            try:
                dt_obj = date(year, month, day)
                d_str  = dt_obj.strftime("%d-%b-%Y")
                d_name = dt_obj.strftime("%A")
            except:
                d_str, d_name = f"Day {day}", ""

            is_holiday    = day in holiday_day_nums
            is_wfh        = day in wfh_dict
            is_auto_sunday = (d_name == "Sunday" and not is_holiday and not is_wfh
                               and day not in punched_days)

            if is_holiday:
                fill_c, note = C_HOLIDAY_BG, f"Holiday (Paid – {decimal_to_hhmm(hrs)} hrs)"
            elif is_wfh:
                info   = wfh_dict[day]
                fill_c = C_WFH_BG
                note   = f"WFH  {info.get('in','?')} → {info.get('out','?')}"
            elif is_auto_sunday:
                fill_c, note = C_HOLIDAY_BG, f"Sunday (Paid – {decimal_to_hhmm(hrs)} hrs)"
            else:
                fill_c, note = C_ALT_ROW, ""

            for col, v in enumerate([d_str, d_name, decimal_to_hhmm(hrs), note], 1):
                c = ws.cell(row=row, column=col, value=v)
                c.fill, c.border, c.alignment, c.font = (
                    make_fill(fill_c), make_border(), make_align(), make_font()
                )
            row += 1

        wk_target    = get_effective_week_target(wk, year, month, current_daily, leave_by_week)
        wk_hrs_dec   = sum_week_hours(week_dict[wk])
        excess       = max(0.0, wk_hrs_dec - wk_target)
        shortage     = max(0.0, wk_target  - wk_hrs_dec)
        summary_fill = make_fill(C_NEUTRAL_BG)

        for label, val in [
            ("Total Worked (Week)", decimal_to_hhmm(wk_hrs_dec)),
            ("Target (Week)",       decimal_to_hhmm(wk_target)),
            ("Excess Hours",        decimal_to_hhmm(excess)),
            ("Shortage Hours",      decimal_to_hhmm(shortage)),
        ]:
            ws.merge_cells(f"A{row}:C{row}")
            for col, v in enumerate([label, None, None, val], 1):
                if col in (2, 3): continue
                c = ws.cell(row=row, column=col, value=v)
                c.font, c.fill, c.border, c.alignment = (
                    make_font(True), summary_fill, make_border(), make_align()
                )
            row += 1
        row += 1

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 26

def write_wfh_sheet(wb, emp_order, raw_records, wfh_records, year, month, period_str):
    ws = wb.create_sheet("WFH Log")

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value, c.font, c.fill, c.alignment = (
        f"Work From Home Log | {period_str}",
        make_font(True, C_HEADER_FG, 14), make_fill(C_HEADER_BG), make_align()
    )

    for col, h in enumerate(["ID", "Employee Name", "Date", "Time In → Out", "Hours"], 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = (
            make_font(True, C_HEADER_FG), make_fill(C_SUBHDR_BG), make_align(), make_border()
        )

    row = 4
    for uid in emp_order:
        wfh_dict = wfh_records.get(uid, {})
        if not wfh_dict:
            continue
        for day in sorted(wfh_dict.keys()):
            info = wfh_dict[day]
            try:
                d_str = date(year, month, day).strftime("%d-%b-%Y (%a)")
            except:
                d_str = f"Day {day}"
            for col, v in enumerate([
                raw_records[uid]['id'],
                raw_records[uid]['name'].title(),
                d_str,
                f"{info.get('in','?')} → {info.get('out','?')}",
                decimal_to_hhmm(info.get('hours', 0.0))
            ], 1):
                c = ws.cell(row=row, column=col, value=v)
                c.fill, c.border, c.alignment, c.font = (
                    make_fill(C_WFH_BG), make_border(), make_align(), make_font()
                )
            row += 1

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 12

def generate_report(employees_dec, emp_order, raw_records, period_str,
                    year, month, daily_target, part_time_list, pt_daily_target,
                    holiday_dates, wfh_records, salary_map=None):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _, row_map = write_raw_data_sheet(wb, employees_dec, emp_order, raw_records,
                                      year, month, daily_target, part_time_list,
                                      pt_daily_target, period_str, holiday_dates, wfh_records)

    _, row_map = write_summary_sheet(wb, employees_dec, emp_order, raw_records,
                                     period_str, year, month, daily_target,
                                     part_time_list, pt_daily_target, row_map,
                                     holiday_dates, wfh_records)

    write_consolidated_sheet(wb, employees_dec, emp_order, raw_records, period_str,
                              year, month, daily_target, part_time_list, pt_daily_target,
                              holiday_dates, wfh_records, row_map, salary_map)

    write_wfh_sheet(wb, emp_order, raw_records, wfh_records, year, month, period_str)

    for uid in emp_order:
        if uid in employees_dec:
            write_individual_sheet(wb, uid, employees_dec[uid], period_str, year, month,
                                   daily_target, uid in part_time_list, pt_daily_target,
                                   holiday_dates, wfh_records, raw_records)

    sheet_order = ["Weekly Summary", "Consolidated Report", "WFH Log", RAW_SHEET]
    for uid in emp_order:
        if uid in employees_dec:
            ws_name = (uid[:28]
                       .replace(":", "").replace("/", "").replace("*", "")
                       .replace("?", "").replace("[", "").replace("]", "").strip())
            sheet_order.append(ws_name)

    existing  = [s.title for s in wb.worksheets]
    ordered   = [s for s in sheet_order if s in existing]
    remaining = [s for s in existing  if s not in ordered]
    for i, name in enumerate(ordered + remaining):
        wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def main():
    st.set_page_config(page_title="Attendance Processor", layout="wide")
    st.title("🕐 Attendance Processor")

    uploaded = st.file_uploader("📂 Upload Attendance XLSX", type=["xlsx"])
    if not uploaded:
        return

    wb_in     = openpyxl.load_workbook(uploaded, read_only=True)
    log_sheet = next((wb_in[n] for n in wb_in.sheetnames if n.lower() == 'logs'), None)
    if not log_sheet:
        st.error("No 'Logs' sheet found.")
        return

    raw_records, emp_order, period_str, year, month = parse_logs_sheet(log_sheet)
    active_employees = [uid for uid in emp_order if raw_records[uid]['punches']]

    for key, default in [
        ('holiday_dates',  []),
        ('wfh_records',    {}),
        ('fixes',          {}),
        ('salary_map',     {}),
        ('part_time_ids',  set()),
    ]:
        if key not in st.session_state:
            st.session_state[key] = default

    with st.sidebar:
        st.header("⚙️ Settings")
        target_weekly = st.number_input(
            "Full-Time Weekly Target (hrs, 7-day week)",
            min_value=1.0, value=59.5, step=0.5
        )
        daily_target = round(target_weekly / 7, 10)

        st.divider()
        st.subheader("🕑 Part-Time Settings")
        pt_daily_target = st.number_input(
            "Part-Time Daily Target (hrs)", min_value=0.5, value=4.0, step=0.5
        )
        st.caption(
            "Upload a reusable Part-Time Master (columns: ID, Name) once — "
            "re-upload the same file each month instead of re-selecting from a list."
        )

        pt_tmpl_col, pt_upl_col = st.columns([1, 2])
        pt_tmpl_col.download_button(
            "📄 Template", make_parttime_template(),
            "part_time_master_template.xlsx", key="pt_template_dl"
        )
        pt_file = pt_upl_col.file_uploader(
            "Upload Part-Time Master", type=["xlsx", "csv"], key="pt_upload"
        )
        if pt_file is not None:
            pt_parsed_ids, pt_parsed_names, pt_skipped = parse_id_name_file(pt_file)
            if pt_parsed_ids:
                st.session_state.part_time_ids.update(pt_parsed_ids)
                st.success(f"✅ Loaded {len(pt_parsed_ids)} part-time employee ID(s).")
                known_ids = {raw_records[u]['id'] for u in active_employees}
                unmatched = [eid for eid in pt_parsed_ids if eid not in known_ids]
                if unmatched:
                    st.warning(
                        "⚠️ These IDs from the part-time file don't match any employee "
                        "in this month's attendance data: "
                        + ", ".join(f"{eid} ({pt_parsed_names.get(eid, '?')})" for eid in unmatched)
                    )
            else:
                st.warning("⚠️ No valid ID/Name rows found in the uploaded file.")

            if pt_skipped:
                with st.expander(f"⚠️ {len(pt_skipped)} row(s) in the file couldn't be read"):
                    for row_num, eid, reason in pt_skipped:
                        st.write(f"Row {row_num} (ID: {eid}): {reason}")

        with st.expander("✏️ Manually set / override part-time employees"):
            for uid in active_employees:
                eid     = raw_records[uid]['id']
                checked = st.checkbox(
                    raw_records[uid]['name'].title(),
                    value=(eid in st.session_state.part_time_ids),
                    key=f"pt_{uid}"
                )
                if checked:
                    st.session_state.part_time_ids.add(eid)
                else:
                    st.session_state.part_time_ids.discard(eid)

        part_time_list = [uid for uid in active_employees
                           if raw_records[uid]['id'] in st.session_state.part_time_ids]

        st.divider()
        st.subheader("🏖️ Office Holidays")
        st.caption("Credited at each employee's daily target (8.30 hrs full-time / part-time rate) as paid holiday, incl. Sundays; not counted as a working day.")

        new_holiday = st.date_input(
            "Pick holiday date",
            value=date(year, month, 1),
            min_value=date(year, month, 1),
            max_value=date(year, month, calendar.monthrange(year, month)[1]),
            key="holiday_picker"
        )
        if st.button("➕ Add Holiday"):
            if new_holiday not in st.session_state.holiday_dates:
                st.session_state.holiday_dates.append(new_holiday)
                st.success(f"Added {new_holiday.strftime('%d-%b-%Y')}")
            else:
                st.warning("Already added.")

        if st.session_state.holiday_dates:
            st.write("**Holidays:**")
            for hd in sorted(st.session_state.holiday_dates):
                c1, c2 = st.columns([3, 1])
                c1.write(hd.strftime("%d-%b-%Y (%a)"))
                if c2.button("✕", key=f"del_hol_{hd}"):
                    st.session_state.holiday_dates.remove(hd)
                    st.rerun()

        st.divider()
        st.subheader("🏠 Work From Home")
        st.caption("Set the employee, date and exact hours worked from home.")

        wfh_emp = st.selectbox(
            "Employee",
            options=active_employees,
            format_func=lambda x: raw_records[x]['name'].title(),
            key="wfh_emp_select"
        ) if active_employees else None

        wfh_date = st.date_input(
            "WFH Date",
            value=date(year, month, 1),
            min_value=date(year, month, 1),
            max_value=date(year, month, calendar.monthrange(year, month)[1]),
            key="wfh_date_picker"
        )

        st.write("**Hours worked from home**")
        ic, oc = st.columns(2)
        with ic:
            st.caption("🟢 Time In")
            wfh_in_h = st.number_input("Hour",   0, 23, 9,  key="wfh_in_h")
            wfh_in_m = st.selectbox("Min", [0, 15, 30, 45],
                                     format_func=lambda x: f"{x:02d}", key="wfh_in_m")
        with oc:
            st.caption("🔴 Time Out")
            wfh_out_h = st.number_input("Hour",   0, 23, 18, key="wfh_out_h")
            wfh_out_m = st.selectbox("Min", [0, 15, 30, 45],
                                      format_func=lambda x: f"{x:02d}", key="wfh_out_m")

        wfh_in_str  = f"{int(wfh_in_h):02d}:{int(wfh_in_m):02d}"
        wfh_out_str = f"{int(wfh_out_h):02d}:{int(wfh_out_m):02d}"
        wfh_hrs, _  = compute_hours_from_pair(wfh_in_str, wfh_out_str)

        if wfh_hrs > 0:
            st.info(f"⏱ {wfh_in_str} → {wfh_out_str} = **{decimal_to_hhmm(wfh_hrs)} hrs**")
        else:
            st.warning("⚠️ Out time must be after In time.")

        if st.button("➕ Add WFH Day", disabled=(wfh_hrs <= 0)):
            if wfh_emp:
                if wfh_emp not in st.session_state.wfh_records:
                    st.session_state.wfh_records[wfh_emp] = {}
                st.session_state.wfh_records[wfh_emp][wfh_date.day] = {
                    'in': wfh_in_str, 'out': wfh_out_str, 'hours': wfh_hrs
                }
                st.success(
                    f"✅ {raw_records[wfh_emp]['name'].title()} | "
                    f"{wfh_date.strftime('%d-%b-%Y')} | "
                    f"{wfh_in_str}→{wfh_out_str} | {decimal_to_hhmm(wfh_hrs)} hrs"
                )

        any_wfh = any(v for v in st.session_state.wfh_records.values())
        if any_wfh:
            st.write("**Current WFH log:**")
            for uid in emp_order:
                wd = st.session_state.wfh_records.get(uid, {})
                if not wd:
                    continue
                st.markdown(f"**{raw_records[uid]['name'].title()}**")
                for d in sorted(wd.keys()):
                    info = wd[d]
                    try:
                        d_lbl = date(year, month, d).strftime("%d-%b (%a)")
                    except:
                        d_lbl = f"Day {d}"
                    cx, cy = st.columns([4, 1])
                    cx.write(
                        f"{d_lbl}  {info.get('in','?')}→{info.get('out','?')}  "
                        f"({decimal_to_hhmm(info.get('hours', 0))} hrs)"
                    )
                    if cy.button("✕", key=f"del_wfh_{uid}_{d}"):
                        del st.session_state.wfh_records[uid][d]
                        st.rerun()

        st.divider()
        st.subheader("💰 Salary Settings")
        st.caption(
            "Upload a reusable Salary Master (columns: ID, Name, Monthly Salary) once — "
            "just re-upload the same file each month instead of retyping salaries. "
            "Per-hour rate = Monthly Salary ÷ Total Target hours; "
            "Calculated Salary = per-hour × Total Hours Worked."
        )

        tmpl_col, upl_col = st.columns([1, 2])
        tmpl_col.download_button(
            "📄 Template", make_salary_template(),
            "salary_master_template.xlsx", key="salary_template_dl"
        )
        salary_file = upl_col.file_uploader(
            "Upload Salary Master", type=["xlsx", "csv"], key="salary_upload"
        )
        if salary_file is not None:
            parsed, parsed_names, skipped_rows = parse_salary_file(salary_file)
            if parsed:
                st.session_state.salary_map.update(parsed)
                st.success(f"✅ Loaded salary for {len(parsed)} employee(s).")
                known_ids = {raw_records[u]['id'] for u in active_employees}
                unmatched = [eid for eid in parsed if eid not in known_ids]
                if unmatched:
                    st.warning(
                        "⚠️ These IDs from the salary file don't match any employee "
                        "in this month's attendance data: "
                        + ", ".join(f"{eid} ({parsed_names.get(eid, '?')})" for eid in unmatched)
                    )
            else:
                st.warning("⚠️ No valid ID/Name/Salary rows found in the uploaded file.")

            if skipped_rows:
                with st.expander(f"⚠️ {len(skipped_rows)} row(s) in the file couldn't be read"):
                    for row_num, eid, reason in skipped_rows:
                        st.write(f"Row {row_num} (ID: {eid}): {reason}")

        with st.expander("✏️ Manually set / override salaries"):
            for uid in active_employees:
                emp_id  = raw_records[uid]['id']
                current = st.session_state.salary_map.get(emp_id, 0.0)
                new_val = st.number_input(
                    raw_records[uid]['name'].title(),
                    min_value=0.0, value=float(current), step=500.0,
                    key=f"salary_{uid}"
                )
                if new_val > 0:
                    st.session_state.salary_map[emp_id] = new_val
                elif emp_id in st.session_state.salary_map:
                    del st.session_state.salary_map[emp_id]

    holiday_dates = st.session_state.holiday_dates
    wfh_records   = st.session_state.wfh_records
    salary_map    = st.session_state.salary_map

    st.header("🔧 Fix Missing Punches")
    any_missing = False
    for uid in active_employees:
        p_dict    = raw_records[uid]['punches']
        emp_fixes = st.session_state.fixes.get(uid, {})
        for day, p in sorted(p_dict.items()):
            if len(p) == 1:
                any_missing = True
                c1, c2, c3, c4 = st.columns([2, 1, 3, 2])
                c1.markdown(f"**{uid}**")
                c2.write(f"Day {day}")
                h = int(p[0].split(':')[0]) if ':' in p[0] else 0
                if h >= 12:
                    c3.warning(f"Out: {p[0]} (In missing)")
                    f_in = c4.text_input("Set In (HH:MM)", value="09:30", key=f"{uid}_{day}_in")
                    try:
                        datetime.strptime(f_in, "%H:%M")
                        emp_fixes[day] = {'in': f_in, 'out': p[0]}
                    except:
                        c4.error("Use HH:MM")
                else:
                    c3.warning(f"In: {p[0]} (Out missing)")
                    f_out = c4.text_input("Set Out (HH:MM)", value="18:00", key=f"{uid}_{day}_out")
                    try:
                        datetime.strptime(f_out, "%H:%M")
                        emp_fixes[day] = {'in': p[0], 'out': f_out}
                    except:
                        c4.error("Use HH:MM")
        if emp_fixes:
            st.session_state.fixes[uid] = emp_fixes

    if not any_missing:
        st.info("✅ No missing punches detected.")

    employees_dec = build_employees_dec(
        active_employees, raw_records, st.session_state.fixes, wfh_records, year, month,
        holiday_dates, part_time_list, pt_daily_target, daily_target
    )

    st.header("📊 Attendance Summary Preview")
    num_sundays  = len(get_month_sundays(year, month))
    num_holidays = len([hd for hd in holiday_dates if hd.year == year and hd.month == month])

    preview = []
    for uid in active_employees:
        if uid not in employees_dec:
            continue
        wd           = wfh_records.get(uid, {})
        wfh_count    = len(wd)
        total_wfh_h  = sum(v.get('hours', 0.0) for v in wd.values())
        days_worked  = get_days_worked(uid, raw_records, wfh_records, holiday_dates, year, month)
        leave_days   = get_leave_days(uid, raw_records, year, month, holiday_dates, wfh_records)
        hol_on_leave = get_holidays_on_leave(uid, raw_records, year, month, holiday_dates, wfh_records)

        week_dict     = employees_dec[uid]
        current_daily = pt_daily_target if uid in part_time_list else daily_target
        leave_by_week = get_leave_days_by_week(uid, raw_records, year, month,
                                                holiday_dates, wfh_records)
        for wk in sorted(week_dict.keys()):
            wk_target     = get_effective_week_target(wk, year, month, current_daily, leave_by_week)
            wk_hrs        = sum_week_hours(week_dict[wk])
            net           = round(wk_hrs - wk_target, 2)
            preview.append({
                "Employee":          raw_records[uid]['name'].title(),
                "Week":              f"Week {wk}",
                "Hrs Worked":        decimal_to_hhmm(wk_hrs),
                "Target":            decimal_to_hhmm(wk_target),
                "Excess":            decimal_to_hhmm(max(0, wk_hrs - wk_target)),
                "Shortage":          decimal_to_hhmm(max(0, wk_target - wk_hrs)),
                "Net Hours":         net,
                "Status":            "Excess" if net > 0 else ("Shortage" if net < 0 else "On Target"),
                "WFH Days":          wfh_count,
                "WFH Hrs":           decimal_to_hhmm(total_wfh_h),
                "Leave Days":        leave_days,
                "Holidays on Leave": hol_on_leave,
                "Sundays":           num_sundays,
                "Holidays":          num_holidays,
            })

    if preview:
        st.dataframe(preview, use_container_width=True)
        st.info(
            "💡 **How editing works in Excel:** Edit the "
            "**'Hours Worked ✏️'** column (col D) in the **Weekly Summary** sheet — "
            "the **Consolidated Report** updates automatically via Excel SUMIF formulas."
        )
    else:
        st.info("No data to preview yet.")

    if salary_map:
        st.header("💰 Salary Preview")
        salary_preview = []
        for uid in active_employees:
            if uid not in employees_dec:
                continue
            emp_id = raw_records[uid]['id']
            monthly_salary = salary_map.get(emp_id)
            if not monthly_salary:
                continue
            current_daily    = pt_daily_target if uid in part_time_list else daily_target
            week_dict        = employees_dec[uid]
            leave_by_week    = get_leave_days_by_week(uid, raw_records, year, month,
                                                        holiday_dates, wfh_records)
            total_hours_dec  = sum(sum(d.values()) for d in week_dict.values())
            total_target_dec = sum(get_effective_week_target(wk, year, month, current_daily, leave_by_week)
                                    for wk in week_dict)
            total_excess_dec = sum(max(0.0, sum(d.values())
                                    - get_effective_week_target(wk, year, month, current_daily, leave_by_week))
                                    for wk, d in week_dict.items())
            capped_hours_dec = total_hours_dec - total_excess_dec
            net              = round(total_hours_dec - total_target_dec, 2)
            per_hour         = monthly_salary / total_target_dec if total_target_dec > 0 else 0
            calc_salary      = round(per_hour * total_hours_dec, 2)
            salary_preview.append({
                "Employee":          raw_records[uid]['name'].title(),
                "Total Hours":       decimal_to_hhmm(capped_hours_dec),
                "Total Target":      decimal_to_hhmm(total_target_dec),
                "Excess":            decimal_to_hhmm(total_excess_dec),
                "Net Hours":         net,
                "Monthly Salary":    monthly_salary,
                "Per-Hour Rate":     round(per_hour, 2),
                "Calculated Salary": calc_salary,
            })
        if salary_preview:
            st.dataframe(salary_preview, use_container_width=True)
        missing = [f"{raw_records[uid]['name'].title()} (ID: {raw_records[uid]['id']})"
                   for uid in active_employees
                   if uid in employees_dec and not salary_map.get(raw_records[uid]['id'])]
        if missing:
            st.warning("⚠️ No salary set for: " + ", ".join(missing))

    st.header("📥 Download Final Report")
    if st.button("Generate Excel Report", type="primary"):
        buf = generate_report(
            employees_dec, active_employees, raw_records, period_str,
            year, month, daily_target, part_time_list, pt_daily_target,
            holiday_dates, wfh_records, salary_map
        )
        st.download_button(
            "⬇️ Download attendance_report.xlsx",
            buf, "attendance_report.xlsx"
        )

if __name__ == "__main__":
    main()
